import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import requests
from openpyxl import Workbook, load_workbook

def salvar_livro():
    isbn = entry_isbn.get()
    nome = entry_nome.get()
    autor = entry_autor.get()
    preco = entry_preco.get()
    quantidade = entry_quantidade.get()

    # Verifica se o campo de nome está vazio
    if not nome:
        messagebox.showerror("Erro", "Por favor, insira o nome do livro.")
        return

    # Verifica se o arquivo já existe ou cria um novo
    try:
        wb = load_workbook("livros.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["ISBN", "Nome", "Autor", "Preço", "Quantidade"])

    # Verifica se o livro já existe na planilha
    livro_existente = False
    row_index = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row, values_only=True), start=2):
        if (row[0] == isbn) or ((row[1] == nome) and (row[2] == autor)):
            row_index = idx
            ws.cell(row=row_index, column=5).value = int(ws.cell(row=row_index, column=5).value) + int(quantidade if quantidade else 1)
            livro_existente = True
            break

    # Se o livro não existir, adiciona-o à planilha
    if not livro_existente:
        row_index = ws.max_row + 1
        ws.append([isbn, nome, autor, preco if preco else None, quantidade if quantidade else 1])

    # Salva as alterações no arquivo
    wb.save("livros.xlsx")
    limpar_campos()
    entry_busca.delete(0, tk.END)  # Limpa o campo de pesquisa
    label_status.config(text="Livro adicionado com sucesso!")
    atualizar_tabela()

def editar_livro():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showerror("Erro", "Selecione um livro para editar.")
        return

    # Obter o índice da linha selecionada na tabela
    row_index = int(table.index(selected_item)) + 1

    # Verifica se o arquivo existe
    try:
        wb = load_workbook("livros.xlsx")
        ws = wb.active
    except FileNotFoundError:
        messagebox.showerror("Erro", "Nenhum livro encontrado.")
        return

    # Obter as informações atuais do livro na planilha
    isbn_atual = ws.cell(row=row_index, column=1).value
    nome_atual = ws.cell(row=row_index, column=2).value
    autor_atual = ws.cell(row=row_index, column=3).value
    preco_atual = ws.cell(row=row_index, column=4).value
    quantidade_atual = ws.cell(row=row_index, column=5).value

    # Obter as novas informações do livro dos campos de entrada
    isbn_novo = entry_isbn.get()
    nome_novo = entry_nome.get()
    autor_novo = entry_autor.get()
    preco_novo = entry_preco.get()
    quantidade_novo = entry_quantidade.get()

    # Atualizar as informações do livro na planilha
    ws.cell(row=row_index, column=1).value = isbn_novo if isbn_novo else isbn_atual
    ws.cell(row=row_index, column=2).value = nome_novo if nome_novo else nome_atual
    ws.cell(row=row_index, column=3).value = autor_novo if autor_novo else autor_atual
    ws.cell(row=row_index, column=4).value = preco_novo if preco_novo else preco_atual
    ws.cell(row=row_index, column=5).value = quantidade_novo if quantidade_novo else quantidade_atual

    # Salvar as alterações no arquivo
    wb.save("livros.xlsx")
    limpar_campos()
    entry_busca.delete(0, tk.END)  # Limpa o campo de pesquisa
    label_status.config(text="Livro editado com sucesso!")
    atualizar_tabela()

def deletar_livro():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showerror("Erro", "Selecione um livro para deletar.")
        return

    # Obter o índice da linha selecionada na tabela
    row_index = int(table.index(selected_item)) + 1

    # Verifica se o arquivo existe
    try:
        wb = load_workbook("livros.xlsx")
        ws = wb.active
    except FileNotFoundError:
        messagebox.showerror("Erro", "Nenhum livro encontrado.")
        return

    # Deletar o livro da planilha
    ws.delete_rows(row_index)

    # Salvar as alterações no arquivo
    wb.save("livros.xlsx")
    limpar_campos()
    entry_busca.delete(0, tk.END)  # Limpa o campo de pesquisa
    label_status.config(text="Livro deletado com sucesso!")
    atualizar_tabela()

def buscar_livro():
    query = entry_busca.get()

    if not query:
        messagebox.showerror("Erro", "Digite um ISBN ou nome de livro para buscar.")
        return

    try:
        response = requests.get(f"https://www.googleapis.com/books/v1/volumes?q={query}")
        data = response.json()

        if "items" in data:
            item = data["items"][0]
            volume_info = item["volumeInfo"]
            isbn = volume_info.get("industryIdentifiers", [{}])[0].get("identifier", "")
            nome = volume_info.get("title", "")
            autor = volume_info.get("authors", [""])[0]

            entry_isbn.delete(0, tk.END)
            entry_nome.delete(0, tk.END)
            entry_autor.delete(0, tk.END)

            entry_isbn.insert(0, isbn)
            entry_nome.insert(0, nome)
            entry_autor.insert(0, autor)
        else:
            messagebox.showerror("Erro", "Livro não encontrado.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao buscar livro: {str(e)}")

def visualizar_livros():
    atualizar_tabela()

def limpar_campos():
    entry_isbn.delete(0, tk.END)
    entry_nome.delete(0, tk.END)
    entry_autor.delete(0, tk.END)
    entry_preco.delete(0, tk.END)
    entry_quantidade.delete(0, tk.END)

def atualizar_tabela():
    try:
        wb = load_workbook("livros.xlsx")
        ws = wb.active

        # Limpar tabela atual
        for row in table.get_children():
            table.delete(row)

        # Atualizar tabela com os livros
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            table.insert("", "end", text=str(i), values=row)

    except FileNotFoundError:
        label_status.config(text="Nenhum livro encontrado.")

root = tk.Tk()
root.title("Gerenciador de Livros")
root_width = root.winfo_screenwidth() // 2
root_height = root.winfo_screenheight() // 2
root.geometry(f"{root_width}x{root_height}")

label_isbn = tk.Label(root, text="ISBN:")
label_isbn.grid(row=0, column=0, padx=10, pady=5)
entry_isbn = tk.Entry(root)
entry_isbn.grid(row=0, column=1, padx=10, pady=5)

label_nome = tk.Label(root, text="Nome:")
label_nome.grid(row=1, column=0, padx=10, pady=5)
entry_nome = tk.Entry(root)
entry_nome.grid(row=1, column=1, padx=10, pady=5)

label_autor = tk.Label(root, text="Autor:")
label_autor.grid(row=2, column=0, padx=10, pady=5)
entry_autor = tk.Entry(root)
entry_autor.grid(row=2, column=1, padx=10, pady=5)

label_preco = tk.Label(root, text="Preço (opcional):")
label_preco.grid(row=3, column=0, padx=10, pady=5)
entry_preco = tk.Entry(root)
entry_preco.grid(row=3, column=1, padx=10, pady=5)

label_quantidade = tk.Label(root, text="Quantidade:")
label_quantidade.grid(row=4, column=0, padx=10, pady=5)
entry_quantidade = tk.Entry(root)
entry_quantidade.grid(row=4, column=1, padx=10, pady=5)

button_adicionar = tk.Button(root, text="Adicionar Livro", command=salvar_livro)
button_adicionar.grid(row=5, column=0, columnspan=2, padx=10, pady=5, sticky="WE")

button_editar = tk.Button(root, text="Editar Livro", command=editar_livro)
button_editar.grid(row=6, column=0, columnspan=2, padx=10, pady=5, sticky="WE")

button_deletar = tk.Button(root, text="Deletar Livro", command=deletar_livro)
button_deletar.grid(row=7, column=0, columnspan=2, padx=10, pady=5, sticky="WE")

label_busca = tk.Label(root, text="Busca por ISBN ou Nome:")
label_busca.grid(row=8, column=0, padx=10, pady=5)
entry_busca = tk.Entry(root)
entry_busca.grid(row=8, column=1, padx=10, pady=5)

button_buscar = tk.Button(root, text="Buscar", command=buscar_livro)
button_buscar.grid(row=8, column=2, padx=10, pady=5)

label_status = tk.Label(root, text="")
label_status.grid(row=9, column=0, columnspan=3, padx=10, pady=5)

# Tabela para visualização dos livros
table_frame = tk.Frame(root)
table_frame.grid(row=0, column=3, rowspan=10, padx=10, pady=5, sticky="NSEW")

table = ttk.Treeview(table_frame, columns=("ISBN", "Nome", "Autor", "Preço", "Quantidade"))
table.heading("#0", text="ID")
table.column("#0", width=30)
table.heading("ISBN", text="ISBN")
table.heading("Nome", text="Nome")
table.heading("Autor", text="Autor")
table.heading("Preço", text="Preço")
table.heading("Quantidade", text="Quantidade")
table.pack(expand=True, fill="both")

# Calcular a largura e a altura mínimas necessárias para exibir todos os elementos sem cortes
min_width = max(1200, root.winfo_reqwidth())
min_height = max(350, root.winfo_reqheight())

# Definir o tamanho da janela
root.minsize(min_width, min_height)


# Definir a largura e a altura máximas da Treeview
max_table_width = int(root_width * 0.9)
max_table_height = int(root_height * 0.9)

# Verificar se a largura e a altura da Treeview excedem os limites máximos
table_width = min(max_table_width, table.winfo_reqwidth())
table_height = min(max_table_height, table.winfo_reqheight())

# Criar a Treeview com o Frame
table_frame = tk.Frame(root)
table_frame.grid(row=0, column=3, rowspan=10, padx=10, pady=5, sticky="NSEW")

# Criar a Treeview
table = ttk.Treeview(table_frame, columns=("ISBN", "Nome", "Autor", "Preço", "Quantidade"), show="headings")
table.grid(row=0, column=0, sticky="NSEW")

# Criar as Scrollbars
vsb = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
vsb.grid(row=0, column=1, sticky="NS")
hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=table.xview)
hsb.grid(row=1, column=0, sticky="EW")

# Configurar as Scrollbars
table.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

# Configurar o redimensionamento das colunas
table.column("#0", width=30, anchor="center")
table.column("ISBN", width=int(table_width * 0.2), anchor="w")  # Alinhado à esquerda
table.column("Nome", width=int(table_width * 0.3), anchor="w")  # Alinhado à esquerda
table.column("Autor", width=int(table_width * 0.2), anchor="w")  # Alinhado à esquerda
table.column("Preço", width=70, anchor="w")  # Largura para 7 dígitos, alinhado à esquerda
table.column("Quantidade", width=70, anchor="w")  # Largura para 7 dígitos, alinhado à esquerda

for col in ("ISBN", "Nome", "Autor", "Preço", "Quantidade"):
    table.heading(col, text=col)

# Adicionar os livros à tabela
atualizar_tabela()

root.mainloop()
